"""
GENERADOR DE AGING COMERCIAL v3.0 - VERSIÓN DEFINITIVA
=====================================================
Sistema automático de análisis de cartera — ComercialAndina S.A.C.

FLUJO:
  SQL Server → Python (aging + scoring) → Excel (3 hojas) → Power BI / n8n

ESTADO MVP (Fase 1):
  - Estado de pago: ESTIMADO (probabilístico por tramo)
  - Se reemplazará en Fase 2 cuando exista tabla Pagos en SQL
  - El scoring es coherente internamente; se recalibrará en Fase 2

LIMITACIONES DOCUMENTADAS:
  - Estado pago simulado → Fase 2 (tabla Pagos)
  - Moneda no diferenciada (PEN/USD) → Fase 3
  - Zona horaria no estandarizada → Fase 3
"""

import pandas as pd
import numpy as np
from datetime import date, timedelta
import pyodbc
import os
import logging
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from scoring_bancario import ScoringBancario
import os; print(f"Corriendo desde: {os.path.abspath(__file__)}")
# ══════════════════════════════════════════════════════════════════
# LOGGING
# ══════════════════════════════════════════════════════════════════

logging.basicConfig(level=logging.INFO, format="%(message)s")
logger = logging.getLogger(__name__)

# ══════════════════════════════════════════════════════════════════
# CONFIG GLOBAL
# ══════════════════════════════════════════════════════════════════

HOY          = date(2026, 3, 24)
CREDITO_DIAS = 30
np.random.seed(99)   # Reproducibilidad mientras estado sea simulado

# ══════════════════════════════════════════════════════════════════
# PASO 1 — CONEXIÓN SQL SERVER
# ══════════════════════════════════════════════════════════════════

try:
    conn = pyodbc.connect(
        "DRIVER={SQL Server};SERVER=localhost;DATABASE=ComercialAndina;Trusted_Connection=yes"
    )
    logger.info("[OK] Conexión SQL Server establecida")
except Exception as e:
    logger.error(f"[ERROR] Conexión: {e}")
    exit()

# ══════════════════════════════════════════════════════════════════
# PASO 2 — EXTRACCIÓN DE DATOS
# ══════════════════════════════════════════════════════════════════

query = """
    SELECT
        cf.id_cliente,
        cf.nombre        AS cliente,
        cf.nro_factura,
        ac.fecha         AS fecha_emision,
        SUM(ad.debe)     AS monto_factura
    FROM Clientes_Facturas cf
    JOIN asientos_cabecera ac ON cf.id_asiento = ac.id_asiento
    JOIN asientos_detalle  ad ON cf.id_asiento = ad.id_asiento
    WHERE ad.codigo_cuenta = '12.13'
      AND ad.debe > 0
    GROUP BY cf.id_cliente, cf.nombre, cf.nro_factura, ac.fecha
    ORDER BY ac.fecha ASC
"""
try:
    df = pd.read_sql(query, conn)
    conn.close()
    logger.info(f"[OK] {len(df):,} filas extraídas de SQL")
except Exception as e:
    logger.error(f"[ERROR] Query SQL: {e}")
    exit()

# ══════════════════════════════════════════════════════════════════
# PASO 3 — LIMPIEZA
# ══════════════════════════════════════════════════════════════════

df["fecha_emision"] = pd.to_datetime(df["fecha_emision"]).dt.date
df = df.drop_duplicates(subset=["nro_factura"])
logger.info(f"[OK] {len(df):,} facturas únicas")

# FIX: Forzar tipo string en id_cliente para evitar type mismatch en merge
df["id_cliente"] = df["id_cliente"].astype(str)

# ══════════════════════════════════════════════════════════════════
# PASO 4 — VENCIMIENTOS Y DÍAS VENCIDOS
# ══════════════════════════════════════════════════════════════════

df["fecha_vencimiento"] = df["fecha_emision"].apply(
    lambda d: d + timedelta(days=CREDITO_DIAS)
)
df["dias_vencidos"] = df["fecha_vencimiento"].apply(
    lambda d: max(0, (HOY - d).days)
)
logger.info("[OK] Vencimientos calculados")

# ══════════════════════════════════════════════════════════════════
# PASO 5 — TRAMO DE AGING
# ══════════════════════════════════════════════════════════════════

conditions = [
    df["dias_vencidos"] <= 30,
    df["dias_vencidos"] <= 60,
    df["dias_vencidos"] <= 90,
]
choices = ["0–30 días", "31–60 días", "61–90 días"]
df["tramo"] = np.select(conditions, choices, default="+90 días")
logger.info("[OK] Tramos de aging asignados")

# ══════════════════════════════════════════════════════════════════
# PASO 6 — ESTADO DE PAGO (FASE 1: ESTIMADO)
# ══════════════════════════════════════════════════════════════════
# FUNCIÓN TEMPORAL — se reemplaza en Fase 2 con JOIN a tabla Pagos real:
#   LEFT JOIN Pagos p ON cf.nro_factura = p.nro_factura
#   → estado = CASE WHEN p.nro_factura IS NOT NULL THEN 'Pagado' ELSE 'Pendiente' END

def estado_pago_fase1(row):
    if row["fecha_emision"] < date(2026, 1, 1):
        return "Pagado"   # Historial 2025 asumido pagado
    prob = {
        "0–30 días":  0.30,
        "31–60 días": 0.50,
        "61–90 días": 0.25,
        "+90 días":   0.10,
    }
    return "Pagado" if np.random.random() < prob.get(row["tramo"], 0.10) else "Pendiente"

df["estado"] = df.apply(estado_pago_fase1, axis=1)
logger.info("[OK] Estado de pago asignado (Fase 1 — estimado)")

# Limpiar tramo para facturas pagadas (no tienen aging relevante)
df.loc[df["estado"] == "Pagado", "tramo"] = None

# ══════════════════════════════════════════════════════════════════
# PASO 7 — SCORING BANCARIO
# ══════════════════════════════════════════════════════════════════

logger.info("\n[INFO] Calculando scoring por cliente...\n")

scoring         = ScoringBancario(fecha_hoy=HOY, debug=True)
scoring_cliente = scoring.aplicar_scoring_dataframe(df)
scoring_cliente["id_cliente"] = scoring_cliente["id_cliente"].astype(str)

df = df.merge(
    scoring_cliente[["id_cliente", "score", "categoria_scoring"]],
    on="id_cliente",
    how="left"
)

# Sanity check — NaN en score indica type mismatch no resuelto
nan_scores = df["score"].isna().sum()
if nan_scores > 0:
    logger.warning(f"⚠️  {nan_scores} filas sin score — revisar tipos de id_cliente")
else:
    logger.info(
        f"\n[OK] Scoring completo: "
        f"min={df['score'].min():.1f} | "
        f"max={df['score'].max():.1f} | "
        f"prom={df['score'].mean():.1f}"
    )

# ══════════════════════════════════════════════════════════════════
# PASO 8 — ACCIÓN SUGERIDA (score + tramo)
# ══════════════════════════════════════════════════════════════════

def asignar_accion(row):
    """
    Matriz de cobranza: combina score de riesgo con días vencidos.
    Pagadas → sin acción.
    Pendientes → acción proporcional al riesgo real.
    """
    if row["estado"] == "Pagado":
        return "—"

    score = row["score"]
    tramo = row["tramo"]

    if score >= 85:
        return "Contacto administrativo"
    elif score >= 70:
        return "Recordatorio amable" if tramo == "0–30 días" else "Recordatorio formal"
    elif score >= 55:
        return "Carta de cobranza" if tramo != "+90 días" else "Gestión activa"
    else:
        return "Aviso prejudicial"

df["accion_sugerida"] = df.apply(asignar_accion, axis=1)
logger.info("[OK] Acciones de cobranza asignadas")

# ══════════════════════════════════════════════════════════════════
# PASO 9 — RESUMEN POR CLIENTE
# ══════════════════════════════════════════════════════════════════

# Columnas auxiliares para evitar closures frágiles en .agg()
df["_es_pagado"]       = df["estado"] == "Pagado"
df["_es_pendiente"]    = df["estado"] == "Pendiente"
df["_monto_pagado"]    = df["monto_factura"].where(df["_es_pagado"],    0)
df["_monto_pendiente"] = df["monto_factura"].where(df["_es_pendiente"], 0)

resumen = df.groupby("id_cliente").agg(
    cliente              = ("cliente",          "first"),
    total_facturas       = ("nro_factura",       "count"),
    total_facturado      = ("monto_factura",     "sum"),
    facturas_pagadas     = ("_es_pagado",        "sum"),
    monto_pagado         = ("_monto_pagado",     "sum"),
    facturas_pendientes  = ("_es_pendiente",     "sum"),
    monto_pendiente      = ("_monto_pendiente",  "sum"),
    facturas_0_30        = ("tramo", lambda x: (x == "0–30 días").sum()),
    facturas_31_60       = ("tramo", lambda x: (x == "31–60 días").sum()),
    facturas_61_90       = ("tramo", lambda x: (x == "61–90 días").sum()),
    facturas_mas_90      = ("tramo", lambda x: (x == "+90 días").sum()),
    dias_max_vencidos    = ("dias_vencidos",     "max"),
    score                = ("score",             "first"),
    categoria_scoring    = ("categoria_scoring", "first"),
).reset_index()

# Limpiar columnas auxiliares
df.drop(columns=["_es_pagado", "_es_pendiente", "_monto_pagado", "_monto_pendiente"], inplace=True)

for col in ["total_facturado", "monto_pagado", "monto_pendiente"]:
    resumen[col] = resumen[col].round(2)

resumen = resumen.sort_values("monto_pendiente", ascending=False)
logger.info(f"[OK] Resumen: {len(resumen)} clientes")

# ══════════════════════════════════════════════════════════════════
# PASO 10 — CARTERA PENDIENTE
# ══════════════════════════════════════════════════════════════════

cartera_pendiente = resumen[resumen["facturas_pendientes"] > 0].copy()
cartera_pendiente = cartera_pendiente.sort_values("monto_pendiente", ascending=False)
logger.info(f"[OK] Cartera pendiente: {len(cartera_pendiente)} clientes con deuda")

# ══════════════════════════════════════════════════════════════════
# PASO 11 — EXPORTAR EXCEL
# ══════════════════════════════════════════════════════════════════

output = "aging_comercial_andina.xlsx"
if os.path.exists(output):
    try:
        os.remove(output)
        logger.info("[OK] Versión anterior eliminada")
    except PermissionError:
        logger.error(f"[ERROR] {output} está abierto en Excel. Ciérralo e intenta de nuevo.")
        exit()

try:
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Aging Detalle",    index=False)
        resumen.to_excel(writer, sheet_name="Resumen General",    index=False)
        cartera_pendiente.to_excel(writer, sheet_name="Cartera Pendiente", index=False)
    logger.info(f"[OK] Excel generado: {output}")
except Exception as e:
    logger.error(f"[ERROR] Excel: {e}")
    exit()

# ══════════════════════════════════════════════════════════════════
# PASO 12 — FORMATO EXCEL
# ══════════════════════════════════════════════════════════════════

wb = load_workbook(output)

def formato_hoja(ws):
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    borde = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )
    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = borde
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 40)

for sheet in ["Aging Detalle", "Resumen General", "Cartera Pendiente"]:
    formato_hoja(wb[sheet])
wb.save(output)
logger.info("[OK] Formateo aplicado")

# ══════════════════════════════════════════════════════════════════
# PASO 13 — REPORTE FINAL EN CONSOLA
# ══════════════════════════════════════════════════════════════════

pendientes       = df[df["estado"] == "Pendiente"]
categorias_count = resumen["categoria_scoring"].value_counts()

print(f"\n{'='*60}")
print(f"  AGING COMERCIAL ANDINA — {HOY.strftime('%d/%m/%Y')}")
print(f"{'='*60}")
print(f"\n  ESTADÍSTICAS GENERALES")
print(f"  Facturas totales      : {len(df):>10,}")
print(f"  Clientes únicos       : {df['cliente'].nunique():>10,}")
print(f"  Monto total facturado : S/ {df['monto_factura'].sum():>14,.2f}")
print(f"\n  COBRANZA")
print(f"  Facturas pagadas      : {(df['estado'] == 'Pagado').sum():>10,}")
print(f"  Facturas pendientes   : {len(pendientes):>10,}")
print(f"  Monto en cobranza     : S/ {pendientes['monto_factura'].sum():>14,.2f}")
print(f"\n  SCORING v3.0")
print(f"  Score promedio        : {resumen['score'].mean():>10.1f}")
print(f"  Score máximo          : {resumen['score'].max():>10.1f}")
print(f"  Score mínimo          : {resumen['score'].min():>10.1f}")
print(f"\n  DISTRIBUCIÓN DE RIESGO")
for cat in ["Excelente", "Bueno", "Aceptable", "Riesgo", "Alto Riesgo"]:
    count = categorias_count.get(cat, 0)
    pct   = (count / len(resumen) * 100) if len(resumen) > 0 else 0
    bar   = "█" * int(pct / 5)
    print(f"  {cat:12} : {count:3} ({pct:5.1f}%)  {bar}")
print(f"\n{'='*60}")
print(f"  Archivo: {output}")
print(f"{'='*60}\n")
import exportar_json as ExJ
ExJ.exportar(df, cartera_pendiente)